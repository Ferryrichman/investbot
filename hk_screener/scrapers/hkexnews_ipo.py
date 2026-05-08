"""
HKEX 官方 New Listing Report (NLR) scraper
數據源:
  Main Board: https://www2.hkexnews.hk/.../New-Listing-Report/Main/NLR{YYYY}_Eng.xlsx
  GEM:        https://www2.hkexnews.hk/.../New-Listing-Report/GEM/e_newlistings{YYYY}.xlsx

Main Board XLSX 結構 (headers row 2):
  Stock Code | Company Name | Date of Prospectus | Date of Listing | Sponsor(s) |
  Reporting Accountants | Valuer(s) | Funds Raised (HK$) | IPO Subscription Price (HK$)
  - 多行 (a/b/c) tier 分開顯示,需要 sum funds_raised 取總

GEM XLSX 結構 (data rows from ~r12):
  col 0: Listing date | col 1: Stock code | col 2: Company | col 5: Offer price
  col 9: Funds raised | col 12: MC at listing | col 14: Industry | col 16: Incorp | col 17: Method
  GEM 冇 sponsor 資料 (要手動補)
"""
import io
import requests
import openpyxl
from collections import defaultdict
from datetime import datetime, timezone

from ..config import IPO_WITHIN_YEARS
from ..db import init_db, get_conn

HEADERS = {"User-Agent": "Mozilla/5.0"}

MAIN_URL = "https://www2.hkexnews.hk/-/media/HKEXnews/Homepage/New-Listings/New-Listing-Information/New-Listing-Report/Main/NLR{year}_Eng.xlsx"
GEM_URL = "https://www2.hkexnews.hk/-/media/HKEXnews/Homepage/New-Listings/New-Listing-Information/New-Listing-Report/GEM/e_newlistings{year}.xlsx"


def _normalize_code(raw):
    """Code might be '02498' or 2498 → return '2498' (strip leading zeros)"""
    if raw is None:
        return None
    s = str(raw).strip().strip('"').lstrip("0")
    if not s:
        return "0"
    return s if s.isdigit() else None


def _to_iso_date(v):
    """Datetime / string → 'YYYY-MM-DD'"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    # Try parse common formats
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    return None


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("HK$", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def parse_main_board_xlsx(content, year):
    """Parse Main Board NLR XLSX,return list of dict per stock"""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    # Headers in row 2; data from row 3
    # Column map (0-indexed): 1=Code, 2=Name, 3=Prospectus, 4=ListingDate, 5=Sponsor, 8=FundsRaised, 9=Price
    aggregated = defaultdict(lambda: {"funds_raised_hkd": 0.0, "rows": []})
    current_code = None
    current_data = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 10:
            continue
        # Tier "(b)" rows have ditto "" in code; use last seen code
        raw_code = row[1]
        if raw_code and str(raw_code).strip() not in ('"', '""', '“', '”'):
            code = _normalize_code(raw_code)
            if code:
                current_code = code
                current_data = {
                    "code": code,
                    "name": str(row[2]).strip() if row[2] else None,
                    "listing_date": _to_iso_date(row[4]),
                    "prospectus_date": _to_iso_date(row[3]),
                    "sponsor": str(row[5]).strip() if row[5] else None,
                    "ipo_price_hkd": _to_float(row[9]),
                    "year": year,
                }
                aggregated[code].update({k: v for k, v in current_data.items() if v is not None})
        # Always sum funds_raised (per tier row)
        if current_code:
            fr = _to_float(row[8])
            if fr:
                aggregated[current_code]["funds_raised_hkd"] += fr
                aggregated[current_code].setdefault("code", current_code)

    # Finalize - drop entries without listing date (might be junk rows)
    out = []
    for code, d in aggregated.items():
        if not d.get("listing_date"):
            continue
        # Clean sponsor: strip newlines, take first if multiple, but keep all comma-separated
        sponsor = d.get("sponsor")
        if sponsor:
            sponsor = " / ".join(s.strip() for s in sponsor.replace("\n", "/").split("/") if s.strip())
        d["sponsor"] = sponsor
        d["raise_amount_hkd"] = d.pop("funds_raised_hkd", 0) or None
        d["board"] = "Main"
        out.append(d)
    return out


def parse_gem_xlsx(content, year):
    """Parse GEM e_newlistings XLSX. No sponsor info available."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    # GEM has multi-line headers; data starts from where col 0 is a date and col 1 is a code
    out = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or len(row) < 10:
            continue
        listing_dt = row[0]
        code = row[1]
        if not isinstance(listing_dt, datetime):
            continue
        norm_code = _normalize_code(code)
        if not norm_code or not norm_code.isdigit():
            continue
        out.append({
            "code": norm_code,
            "name": str(row[2]).strip() if row[2] else None,
            "listing_date": listing_dt.date().isoformat(),
            "ipo_price_hkd": _to_float(row[5]),
            "raise_amount_hkd": _to_float(row[9]),
            "sponsor": None,  # GEM XLSX 冇 sponsor
            "board": "GEM",
            "year": year,
        })
    return out


def fetch_year(year, board="Main"):
    """Download + parse one year's XLSX"""
    url = (MAIN_URL if board == "Main" else GEM_URL).format(year=year)
    try:
        r = requests.get(url, headers=HEADERS, timeout=60)
        if r.status_code != 200 or len(r.content) < 5000:
            print(f"  [{board} {year}] HTTP {r.status_code}, size={len(r.content)} - skip")
            return []
        if board == "Main":
            return parse_main_board_xlsx(r.content, year)
        else:
            return parse_gem_xlsx(r.content, year)
    except Exception as e:
        print(f"  [{board} {year}] error: {e}")
        return []


def run(years=None):
    """
    Main entry — fetch all years' NLR + GEM XLSX, write to ipo_info table
    Only updates rows where raise_amount or sponsor is currently NULL (preserve manual overrides)
    """
    init_db()
    if years is None:
        current = datetime.now().year
        # Fetch IPO_WITHIN_YEARS + 1 buffer (for sponsor 5-yr lookback we want more)
        years = list(range(current - max(IPO_WITHIN_YEARS, 6), current + 1))

    all_records = []
    for year in years:
        for board in ("Main", "GEM"):
            recs = fetch_year(year, board)
            print(f"  [{board} {year}] {len(recs)} IPOs")
            all_records.extend(recs)

    print(f"[hkexnews_ipo] Total {len(all_records)} IPO records to write")

    # Write to DB - upsert ipo_info; only fill in raise/sponsor where they're NULL
    now = datetime.now(timezone.utc).isoformat()
    with get_conn() as conn:
        # Get codes that exist in stocks table
        existing_codes = {r[0] for r in conn.execute("SELECT code FROM stocks").fetchall()}

        updated = 0
        new_inserts = 0
        not_in_universe = 0

        for r in all_records:
            code = r["code"]
            if code not in existing_codes:
                not_in_universe += 1
                continue

            # Check existing record
            row = conn.execute(
                "SELECT raise_amount_hkd, sponsor, ipo_price_hkd FROM ipo_info WHERE code = ?",
                (code,)
            ).fetchone()

            if row:
                # Update only NULL fields (preserve manual / Yahoo data)
                cur_raise, cur_sponsor, cur_price = row
                new_raise = cur_raise if cur_raise is not None else r.get("raise_amount_hkd")
                new_sponsor = cur_sponsor if cur_sponsor else r.get("sponsor")
                new_price = cur_price if cur_price is not None else r.get("ipo_price_hkd")
                conn.execute("""
                    UPDATE ipo_info SET
                        listing_date = COALESCE(listing_date, ?),
                        ipo_price_hkd = ?,
                        raise_amount_hkd = ?,
                        sponsor = ?,
                        last_updated = ?
                    WHERE code = ?
                """, (r["listing_date"], new_price, new_raise, new_sponsor, now, code))
                updated += 1
            else:
                conn.execute("""
                    INSERT INTO ipo_info (code, listing_date, ipo_price_hkd, raise_amount_hkd,
                                          sponsor, last_updated)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (code, r["listing_date"], r.get("ipo_price_hkd"),
                      r.get("raise_amount_hkd"), r.get("sponsor"), now))
                new_inserts += 1

    print(f"[hkexnews_ipo] Done: {updated} updated, {new_inserts} new, {not_in_universe} not in universe")
    return updated + new_inserts


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        years = [int(y) for y in sys.argv[1:]]
        run(years=years)
    else:
        run()
